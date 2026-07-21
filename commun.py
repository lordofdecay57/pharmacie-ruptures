# -*- coding: utf-8 -*-
"""Fonctions PURES partagées entre les deux modules métier de l'application.

Ce module ne connaît ni les ruptures GPNC/UNIPHARMA (moteur_ruptures.py) ni
la politique de stock min/max (stock_rotation.py) : il fournit uniquement
les briques génériques que les deux domaines utilisent — parsing, lecture
de fichiers fournisseurs, statistiques de consommation.

C'est le mécanisme de MUTUALISATION voulu par l'architecture : les deux
modules métier importent d'ici ce dont ils ont besoin, mais n'importent
JAMAIS l'un de l'autre. Aucun couplage fonctionnel entre eux.
"""

from __future__ import annotations

import io
import math
import re
import unicodedata
from datetime import date, datetime
from typing import Optional

import pandas as pd

# ---------------------------------------------------------------------------
# Constantes de calcul partagées
# ---------------------------------------------------------------------------

JOURS_PAR_MOIS = 30               # convention rotation mensuelle → journalière
ALPHA_LISSAGE = 0.4               # lissage exponentiel (mois récent : 40 %)
SEUIL_TENDANCE = 0.20             # ±20 % entre récent et référence → ↗ / ↘
SEUILS_VARIABILITE = (0.3, 0.7)   # CV : < 0,3 stable · < 0,7 variable · sinon forte


# ---------------------------------------------------------------------------
# Normalisation / parsing
# ---------------------------------------------------------------------------

def normaliser_libelle(libelle) -> str:
    """Majuscules, sans accents, ponctuation → espace, espaces réduits."""
    if libelle is None or (isinstance(libelle, float) and math.isnan(libelle)):
        return ""
    s = str(libelle)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normaliser_cip(cip) -> str:
    """Ne garde que les chiffres (gère les CIP lus en float : '3400930.0').

    Un CIP « 0 » (placeholder fréquent dans les exports) est traité comme
    absent — sinon deux produits distincts à CIP 0 se rapprocheraient à tort.
    """
    if cip is None or (isinstance(cip, float) and math.isnan(cip)):
        return ""
    s = str(cip).strip()
    if re.fullmatch(r"\d+\.0+", s):  # float Excel → entier
        s = s.split(".")[0]
    s = re.sub(r"\D", "", s)
    return "" if s.strip("0") == "" else s


def variantes_cip(cip: str) -> list:
    """Formes équivalentes d'un CIP pour le matching inter-fichiers.

    Les exports mélangent CIP13 et CIP7 : le CIP13 médicament français
    (13 chiffres, préfixe 3400) contient le CIP7 en positions 6-12
    (ex. 3400932300778 → 3230077, Titanoréine). On rapproche donc les deux
    formes. Les autres EAN13 (parapharmacie…) restent tels quels.
    """
    if not cip:
        return []
    formes = [cip]
    if len(cip) == 13 and cip.startswith("3400"):
        formes.append(cip[5:12])  # CIP7 embarqué dans le CIP13
    return formes


def parser_nombre(val) -> float:
    """Nombre robuste : virgule décimale française, espaces, vide → 0."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if (isinstance(val, float) and math.isnan(val)) else float(val)
    s = str(val).strip().replace(" ", "").replace(" ", "")
    if not s:
        return 0.0
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parser_date(val) -> Optional[date]:
    """Date robuste (formats français en priorité). None si illisible/vide."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in {"nan", "nat", "-", "?"}:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y",
                "%d/%m", "%d-%m"):
        try:
            d = datetime.strptime(s, fmt).date()
            if fmt in ("%d/%m", "%d-%m"):  # jour/mois sans année → année en cours
                d = d.replace(year=date.today().year)
            return d
        except ValueError:
            continue
    try:  # dernier recours : pandas, convention jour d'abord (français)
        d = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return None if pd.isna(d) else d.date()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Chargement des fichiers (.xlsx / .xls / .csv / .pdf)
# ---------------------------------------------------------------------------

def charger_fichier(contenu, nom_fichier: str) -> pd.DataFrame:
    """Charge un fichier Excel/CSV/PDF en DataFrame (colonnes en str).

    ``contenu`` : bytes, chemin, ou objet fichier (upload Streamlit).
    Lève ValueError avec un message clair si le format n'est pas géré.
    """
    nom = (nom_fichier or "").lower()
    if isinstance(contenu, (str,)):  # chemin sur disque
        with open(contenu, "rb") as f:
            data = f.read()
    elif isinstance(contenu, bytes):
        data = contenu
    else:  # objet fichier (BytesIO, UploadedFile…)
        data = contenu.read()

    if nom.endswith(".csv") or nom.endswith(".txt"):
        for encodage in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                texte = data.decode(encodage)
            except UnicodeDecodeError:
                continue
            df = _parser_cadencier_winpharma_csv(texte)
            if df is not None:
                break
            try:
                sep = ";" if texte.count(";") >= texte.count(",") else ","
                df = pd.read_csv(io.StringIO(texte), sep=sep)
                break
            except pd.errors.ParserError:
                continue
        else:
            raise ValueError(f"CSV illisible : {nom_fichier}")
    elif nom.endswith(".xlsx") or nom.endswith(".xlsm"):
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
    elif nom.endswith(".xls"):
        df = pd.read_excel(io.BytesIO(data))  # xlrd requis pour les vieux .xls
    elif nom.endswith(".pdf"):
        df = _charger_pdf(data, nom_fichier)
    else:
        raise ValueError(
            f"Format non géré : {nom_fichier} (attendu .xlsx, .xls, .csv ou .pdf)")

    df.columns = [str(c).strip() for c in df.columns]
    return df


def _charger_pdf(data: bytes, nom_fichier: str) -> pd.DataFrame:
    """Extrait le tableau d'un PDF multi-pages (export type cadencier).

    Le cadencier WinPharma (en-tête « Codes produit ») est reconnu et passe
    par un parseur dédié. Sinon : la première ligne non vide sert d'en-tête,
    l'en-tête répété en haut de chaque page est éliminé. PDF sans traits de
    tableau : repli sur l'alignement du texte. PDF scanné (image) : message
    clair, préférer l'export Excel/CSV.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ValueError("Lecture PDF indisponible : lancez "
                         "« pip install pdfplumber » puis réessayez.")

    brutes: list = []
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:  # pas de traits → colonnes par alignement texte
                    table = page.extract_table({"vertical_strategy": "text",
                                                "horizontal_strategy": "text"})
                    tables = [table] if table else []
                for table in tables:
                    for brute in table:
                        valeurs = ["" if v is None else str(v).strip()
                                   for v in (brute or [])]
                        if any(valeurs):
                            brutes.append(valeurs)
    except Exception:
        raise ValueError(f"PDF illisible : {nom_fichier}")

    if not brutes:
        raise ValueError(
            f"Aucun tableau lisible dans {nom_fichier} — s'il s'agit d'un "
            "PDF scanné (image), préférez un export Excel ou CSV.")

    if any("Codes produit" in " ".join(r) for r in brutes[:40]):
        return _parser_cadencier_winpharma(brutes)

    en_tete, lignes = None, []
    for valeurs in brutes:
        if en_tete is None:
            en_tete = valeurs
        elif valeurs != en_tete:  # ignore l'en-tête répété à chaque page
            lignes.append(valeurs)
    if en_tete is None or not lignes:
        raise ValueError(
            f"Aucun tableau lisible dans {nom_fichier} — s'il s'agit d'un "
            "PDF scanné (image), préférez un export Excel ou CSV.")
    largeur = len(en_tete)  # aligne les lignes incomplètes sur l'en-tête
    lignes = [l[:largeur] + [""] * (largeur - len(l)) for l in lignes]
    return pd.DataFrame(lignes, columns=en_tete)


def _separer_mois_et_total(nombres: list) -> list:
    """12 valeurs au lieu de 13 : le mois le plus ancien et le Total sont
    probablement collés. Sépare le dernier jeton en (mois, total) en validant
    par la somme : Total == somme des 12 mois. Inchangé si aucune coupe ne
    se vérifie (vraie ligne à 12 valeurs sans total, ou décimales).
    """
    fusion = nombres[-1]
    if not fusion.isdigit() or len(fusion) < 2:
        return nombres
    somme_connue = int(round(sum(parser_nombre(v) for v in nombres[:11])))
    for coupe in range(1, len(fusion)):
        mois, total = fusion[:coupe], fusion[coupe:]
        if int(total) == somme_connue + int(mois):
            return nombres[:11] + [mois]
    return nombres


def _parser_cadencier_winpharma(brutes: list) -> pd.DataFrame:
    """Cadencier de stock WinPharma (PDF) → DataFrame prêt pour l'analyse.

    Format observé (4 cellules par ligne) :
      - « Codes produit » : CIP7 et CIP13 empilés (ou EAN seul) ;
      - « Nom / Formes & presentations » : libellé, parfois sur 2 lignes ;
      - « Stock » : quantité en rayon ;
      - achats/ventes : « A … » puis « V … » sur la même cellule, 12 valeurs
        mensuelles en ordre ANTI-chronologique (mois récent en premier) +
        total. On ne garde que la ligne V, remise en ordre chronologique.
    Le bandeau de la pharmacie, les en-têtes répétés par page et la ligne de
    totaux finale sont éliminés.
    """
    mois = None
    for r in brutes:
        joint = " ".join(r)
        if "Codes produit" in joint:
            trouve = re.search(r"((?:[A-Za-zéû]{3}\s+){11}[A-Za-zéû]{3})\s+Total",
                               joint)
            if trouve:
                mois = trouve.group(1).split()
            break
    if not mois:
        raise ValueError("Cadencier WinPharma : en-tête des mois introuvable.")
    colonnes_ventes = [f"Ventes {m}" for m in reversed(mois)]  # récent en DERNIER

    produits = []
    for r in brutes:
        if len(r) < 4:
            continue
        codes_cell, nom, stock_cell, achats_ventes = r[0], r[1], r[2], r[3]
        if "Codes produit" in codes_cell or "CADENCIER" in " ".join(r).upper():
            continue
        codes = re.findall(r"\d{6,}", codes_cell)
        if not codes:  # bandeau de page, ligne de totaux (« Manque: … »)…
            continue
        cip = next((c for c in codes if len(c) == 13), codes[0])
        ventes_v = re.search(r"(?:^|\n)V\s+([\d\s,.]+)", achats_ventes or "")
        ventes: list = []
        if ventes_v:
            nombres = ventes_v.group(1).split()
            if len(nombres) > 12:
                nombres = nombres[:12]  # sans la colonne Total
            elif len(nombres) == 12:
                # Gros vendeurs : le 12e mois et le Total sont COLLÉS dans le
                # PDF (colonne étroite) — ex. « 157218268 » = 1572 + 18268.
                nombres = _separer_mois_et_total(nombres)
            ventes = list(reversed(nombres))  # → ordre chronologique
        ligne = {"Produit": " ".join((nom or "").split()), "CIP": cip,
                 "Stock": stock_cell}
        manquants = len(colonnes_ventes) - len(ventes)
        ventes = ["0"] * max(0, manquants) + ventes  # mois absents = 0 vente
        for colonne, valeur in zip(colonnes_ventes, ventes):
            ligne[colonne] = valeur
        produits.append(ligne)
    if not produits:
        raise ValueError("Cadencier WinPharma : aucune ligne produit lisible.")
    return pd.DataFrame(produits,
                        columns=["Produit", "CIP", "Stock"] + colonnes_ventes)


def _parser_cadencier_winpharma_csv(texte: str) -> Optional[pd.DataFrame]:
    """Cadencier de stock WinPharma exporté en CSV → même format normalisé
    que le parseur PDF : ``Produit`` / ``CIP`` / ``Stock`` / ``Ventes <mois>``
    en ordre CHRONOLOGIQUE.

    Format observé : bandeau de la pharmacie sur les premières lignes, puis
    en-tête ``CIP;Code13Réf;Nom;"Formes & presentations";Stock`` suivi de
    12 mois d'achats « (A) » et 12 mois de ventes « (V) » en ordre
    ANTI-chronologique, et des colonnes « Total ». Les achats et totaux sont
    ignorés, la ligne de totaux finale (sans code produit) est éliminée, et
    le Code13Réf (13 chiffres) est préféré au CIP court pour les appariements.
    Renvoie None si le texte n'est pas un cadencier WinPharma.
    """
    lignes = texte.splitlines()
    entete = next((i for i, l in enumerate(lignes[:40])
                   if l.split(";")[0].strip().strip('"').upper() == "CIP"
                   and "(V)" in l), None)
    if entete is None:
        return None
    brut = pd.read_csv(io.StringIO("\n".join(lignes[entete:])),
                       sep=";", dtype=str)
    brut.columns = [str(c).strip() for c in brut.columns]

    ventes_anti_chrono = [c for c in brut.columns
                          if re.fullmatch(r"(?!Total).+\(V\)", c)]
    if not ventes_anti_chrono:
        return None
    df = pd.DataFrame()
    df["Produit"] = brut.get("Nom", "").fillna("").map(
        lambda s: " ".join(str(s).split()))
    col_code13 = next((c for c in brut.columns
                       if "code13" in _sans_accents(str(c))), None)
    code13 = brut[col_code13] if col_code13 else None
    cip = brut.get("CIP")

    def _meilleur_code(i):
        codes = []
        for serie in (code13, cip):
            if serie is not None:
                codes.append(re.sub(r"\D", "", str(serie.iloc[i] or "")))
        codes = [c for c in codes if len(c) >= 6]
        if not codes:
            return ""
        return next((c for c in codes if len(c) == 13), codes[0])

    df["CIP"] = [_meilleur_code(i) for i in range(len(brut))]
    df["Stock"] = brut.get("Stock", "0").fillna("0")
    for colonne in reversed(ventes_anti_chrono):  # → ordre chronologique
        mois = colonne.replace("(V)", "").strip()
        df[f"Ventes {mois}"] = brut[colonne].fillna("0")
    # Élimine la ligne de totaux (« Qte : 3621 » / « Manque : -8 ») mais garde
    # les produits sans code (parapharmacie) : l'appariement retombe alors
    # sur le libellé.
    totaux = df["Produit"].str.match(r"(?i)\s*(qte|manque)\s*:")
    df = df[((df["CIP"] != "") | (df["Produit"] != "")) & ~totaux]
    df = df.reset_index(drop=True)
    if df.empty:
        return None
    return df


# ---------------------------------------------------------------------------
# Détection automatique des colonnes (proposition, à confirmer dans l'UI)
# ---------------------------------------------------------------------------

_MOTS_CLES = {
    "libelle": ["libell", "produit", "design", "article", "nom", "denomination"],
    "cip": ["cip", "code produit", "code article", "ean", "acl"],
    "stock": ["stock", "qte dispo", "quantite dispo", "disponible"],
    "date_reappro": ["reappro", "réappro", "reapprovisionnement", "retour",
                     "dispo le", "date"],
    "conditionnement": ["conditionnement", "colisage", "pcb", "unite de vente"],
    "commande_en_cours": ["commande en cours", "qte commandee", "cde en cours",
                          "en commande"],
    "peremption": ["peremption", "dluo", "date de peremption", "date limite"],
}


def _sans_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c)).lower()


def detecter_colonne(colonnes, role: str) -> Optional[str]:
    """Propose la colonne la plus probable pour un rôle donné (ou None)."""
    for mot in _MOTS_CLES.get(role, []):
        for col in colonnes:
            if _sans_accents(mot) in _sans_accents(str(col)):
                return col
    return None


def detecter_colonnes_ventes(colonnes) -> list:
    """Propose les colonnes de ventes mensuelles (mois ou mot-clé « vente »)."""
    mois = ["janv", "fevr", "mars", "avr", "mai", "juin", "juil", "aout",
            "sept", "oct", "nov", "dec"]
    trouvees = []
    for col in colonnes:
        c = _sans_accents(str(col))
        if "vente" in c or "sortie" in c or any(m in c for m in mois):
            trouvees.append(col)
    return trouvees


# ---------------------------------------------------------------------------
# Calculs de consommation partagés — utilisés par les DEUX modules métier
# ---------------------------------------------------------------------------

def _depuis_premiere_vente(valeurs: list) -> list:
    """Tronque les zéros de TÊTE d'une série mensuelle chronologique.

    Un produit référencé en cours de période affiche des mois à 0 AVANT sa
    première vente : ce n'est pas de la demande nulle, c'est « pas encore au
    catalogue ». Les inclure dans une moyenne divise la rotation des produits
    récents (génériques nouvellement référencés notamment) et fait
    sous-dimensionner leur stock. Les statistiques se calculent donc depuis
    la première vente. Série entièrement nulle : inchangée.
    """
    premiere = next((i for i, v in enumerate(valeurs) if v != 0), None)
    return valeurs if premiere is None else valeurs[premiere:]


def calculer_rotation_mensuelle(ventes: list, periode: str = "annuelle") -> float:
    """Rotation mensuelle estimée.

    ``ventes`` : valeurs mensuelles en ordre CHRONOLOGIQUE (la plus récente en
    dernier). Les mois à 0 AVANT la première vente (produit pas encore
    référencé) sont exclus du calcul. ``periode`` :
      - "annuelle" : moyenne de toutes les valeurs ;
      - "6mois"    : moyenne des 6 dernières ;
      - "3mois"    : moyenne des 3 dernières ;
      - "1mois"    : le dernier mois seul — le plus réactif, mais sensible à
        un mois atypique (pic, creux ou rupture ponctuelle) ;
      - "lissee"   : lissage exponentiel (α = 0,4) — réactif aux tendances
        récentes, à la hausse COMME à la baisse, sans sur-réagir à un mois
        isolé. Recommandé pour l'analyse quotidienne.
    """
    valeurs = _depuis_premiere_vente([parser_nombre(v) for v in ventes])
    if not valeurs:
        return 0.0
    if periode == "1mois":
        return valeurs[-1]
    if periode == "6mois":
        valeurs = valeurs[-6:]
    elif periode == "3mois":
        valeurs = valeurs[-3:]
    elif periode == "lissee":
        lisse = valeurs[0]
        for v in valeurs[1:]:
            lisse = ALPHA_LISSAGE * v + (1 - ALPHA_LISSAGE) * lisse
        return lisse
    return sum(valeurs) / len(valeurs)


def corriger_faux_zeros(ventes: list):
    """Corrige les mois à 0 vente ENCADRÉS de mois actifs.

    Un 0 entre deux mois vendeurs signifie « produit en rupture », pas
    « personne n'en voulait » : le laisser écrase la rotation et fait
    SOUS-commander précisément les produits qui ont déjà manqué. Les zéros
    sont remplacés par interpolation linéaire entre les mois actifs qui les
    encadrent. Les zéros en DÉBUT ou FIN de période sont conservés
    (lancement, arrêt de commercialisation, rupture en cours).

    Renvoie ``(ventes_corrigees, nb_mois_corriges)``.
    """
    valeurs = [parser_nombre(v) for v in ventes]
    corrigees, nb = list(valeurs), 0
    i = 0
    while i < len(valeurs):
        if valeurs[i] == 0:
            j = i
            while j < len(valeurs) and valeurs[j] == 0:
                j += 1
            if 0 < i and j < len(valeurs):  # encadré de mois actifs
                gauche, droite = valeurs[i - 1], valeurs[j]
                for k in range(j - i):
                    corrigees[i + k] = (gauche + (droite - gauche)
                                        * (k + 1) / (j - i + 1))
                    nb += 1
            i = j
        else:
            i += 1
    return corrigees, nb


def calculer_stock_jours(stock_actuel: float, rotation_mensuelle: float) -> float:
    """Couverture actuelle en jours. Stock 0 → 0 ; rotation nulle → +inf."""
    if stock_actuel <= 0:
        return 0.0
    rotation_journaliere = rotation_mensuelle / JOURS_PAR_MOIS
    if rotation_journaliere <= 0:
        return math.inf
    return stock_actuel / rotation_journaliere


def calculer_tendance(ventes: list, seuil: float = SEUIL_TENDANCE) -> str:
    """Tendance de la demande, en ordre chronologique (récent en dernier).

    - ≥ 4 mois de recul : moyenne des 3 derniers mois vs moyenne globale ;
    - 2-3 mois (cadenciers courts, très fréquents) : dernier mois vs moyenne
      des mois précédents — moins robuste mais la colonne reste vivante ;
    - < 2 mois ou demande nulle : « → stable » (rien à comparer).
    Renvoie « ↗ hausse » / « ↘ baisse » / « → stable » (seuil ±20 %).
    """
    if len(ventes) < 2:
        return "→ stable"
    if len(ventes) >= 4:
        reference = calculer_rotation_mensuelle(ventes, "annuelle")
        recente = calculer_rotation_mensuelle(ventes, "3mois")
    else:
        reference = calculer_rotation_mensuelle(ventes[:-1], "annuelle")
        recente = parser_nombre(ventes[-1])
    if reference <= 0:
        return "→ stable"
    ecart = (recente - reference) / reference
    if ecart >= seuil:
        return "↗ hausse"
    if ecart <= -seuil:
        return "↘ baisse"
    return "→ stable"


def classer_abc(volumes: list) -> list:
    """Classement ABC (loi de Pareto) sur les volumes de ventes.

    A = les plus gros vendeurs jusqu'à 80 % du volume cumulé, B = 80-95 %,
    C = le reste (dont les volumes nuls). Renvoie les classes dans l'ordre
    d'entrée. Le plus gros vendeur est toujours A (cumul évalué AVANT lui).
    """
    valeurs = [parser_nombre(v) for v in volumes]
    total = sum(v for v in valeurs if v > 0)
    classes = ["C"] * len(valeurs)
    if total <= 0:
        return classes
    ordre = sorted(range(len(valeurs)), key=lambda i: valeurs[i], reverse=True)
    cumul = 0.0
    for i in ordre:
        if valeurs[i] <= 0:
            continue
        part_avant = cumul / total
        classes[i] = ("A" if part_avant < 0.80
                      else "B" if part_avant < 0.95 else "C")
        cumul += valeurs[i]
    return classes


def coefficient_variation(ventes: list) -> Optional[float]:
    """Coefficient de variation σ/μ de la demande (None si non calculable).

    Mesure la RÉGULARITÉ des ventes : proche de 0 = très régulier, > 1 =
    très erratique. Le recul se compte depuis la PREMIÈRE vente (les mois
    d'avant référencement gonfleraient artificiellement le CV). Moins de
    3 mois de recul ou demande nulle → None (inconnu).
    """
    valeurs = _depuis_premiere_vente([parser_nombre(v) for v in ventes])
    if len(valeurs) < 3:
        return None
    moyenne = sum(valeurs) / len(valeurs)
    if moyenne <= 0:
        return None
    ecart_type = (sum((v - moyenne) ** 2 for v in valeurs) / len(valeurs)) ** 0.5
    return ecart_type / moyenne


def variabilite_demande(ventes: list) -> str:
    """Variabilité de la demande (coefficient de variation σ/μ), en libellé.

    Sert de base à un stock de sécurité différencié : un produit
    « forte variabilité » mérite plus de marge qu'un produit stable à
    volume égal. Moins de 3 mois de recul ou demande nulle → « » (inconnu).
    """
    cv = coefficient_variation(ventes)
    if cv is None:
        return ""
    if cv < SEUILS_VARIABILITE[0]:
        return f"🟢 stable (CV {cv:.0%})"
    if cv < SEUILS_VARIABILITE[1]:
        return f"🟡 variable (CV {cv:.0%})"
    return f"🔴 forte (CV {cv:.0%})"


def pic_saisonnier(ventes: list, noms_mois: list) -> str:
    """Signale un pic saisonnier probable : un mois ≥ 2× la moyenne.

    Nécessite au moins 6 mois de recul DEPUIS LA PREMIÈRE VENTE pour
    distinguer saison et hasard (les mois d'avant référencement écrasent la
    moyenne et créent de faux pics).
    ``noms_mois`` : libellés alignés sur ``ventes`` (pour nommer le pic).
    """
    completes = [parser_nombre(v) for v in ventes]
    valeurs = _depuis_premiere_vente(completes)
    decalage = len(completes) - len(valeurs)
    if len(valeurs) < 6:
        return ""
    moyenne = sum(valeurs) / len(valeurs)
    if moyenne <= 0:
        return ""
    maximum = max(valeurs)
    if maximum < 2 * moyenne:
        return ""
    nom = ""
    if noms_mois and len(noms_mois) == len(completes):
        nom = str(noms_mois[decalage + valeurs.index(maximum)]
                  ).replace("Ventes", "").strip()
    return f"📈 pic {nom}".strip()


# ---------------------------------------------------------------------------
# Export Excel (mise en forme commune aux deux modules)
# ---------------------------------------------------------------------------

def exporter_classeur(onglets: list, couleurs_par_colonne: Optional[dict] = None
                      ) -> bytes:
    """Classeur Excel commun : en-têtes gras figés, largeurs auto, lignes
    teintées selon ``couleurs_par_colonne`` (ex. {"Urgence": {...}}) quand la
    colonne existe dans l'onglet.

    ``onglets`` : liste de tuples (nom_feuille, DataFrame).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    couleurs_par_colonne = couleurs_par_colonne or {}
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nom, df in onglets:
            df.to_excel(writer, sheet_name=nom, index=False)
            ws = writer.sheets[nom]
            ws.freeze_panes = "A2"
            for cellule in ws[1]:  # en-tête gras sur fond gris clair
                cellule.font = Font(bold=True)
                cellule.fill = PatternFill("solid", fgColor="D9D9D9")
                cellule.alignment = Alignment(vertical="center")
            for j, col in enumerate(df.columns, start=1):  # largeurs auto
                largeur = max([len(str(col))] +
                              [len(str(v)) for v in df[col].head(200)] or [10])
                ws.column_dimensions[get_column_letter(j)].width = min(largeur + 3, 45)
            for colonne, couleurs in couleurs_par_colonne.items():
                if colonne in df.columns:  # code couleur des lignes
                    for i, valeur in enumerate(df[colonne], start=2):
                        couleur = couleurs.get(valeur)
                        if couleur:
                            for cellule in ws[i]:
                                cellule.fill = PatternFill("solid", fgColor=couleur)
    return buffer.getvalue()


def nom_fichier_export(prefixe: str, date_analyse: date) -> str:
    """Nom conventionnel d'un fichier généré : ``prefixe_AAAA-MM-JJ.xlsx``."""
    return f"{prefixe}_{date_analyse:%Y-%m-%d}.xlsx"
