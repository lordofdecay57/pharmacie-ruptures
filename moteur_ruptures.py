# -*- coding: utf-8 -*-
"""Moteur métier de gestion des ruptures de stock (pharmacie).

Module Python PUR, sans interface : toute la logique de calcul vit ici et est
testable indépendamment (voir tests/test_moteur.py). L'interface Streamlit
(app.py) ne fait qu'appeler ce module.

Logique métier (STRICTE, sans buffer — corrigée avec l'utilisateur) :
  1. Périmètre : produits en rupture GPNC ET vendus (rotation > 0 au cadencier).
  2. stock_jours = stock_actuel / (rotation_mensuelle / 30) ; stock 0 → 0 jour.
  3. Apparition :
       - avec date de réappro : stock_jours < jours_avant_reappro (STRICT) ;
       - sans date : stock_jours < 30 (objectif 30 jours de couverture).
     Ex. Titanoréine : réappro 16 j, stock 18 j → 18 ≥ 16 → n'apparaît PAS.
  4. Dépannage : absent des ruptures UNIPHARMA → Onglet 1 (commander) ;
     présent → Onglet 2 (aucune solution).
  5. Cmd = arrondi_sup(rotation_journaliere × couverture_cible − stock), min 1,
     arrondi au conditionnement si l'info existe.
  6. Urgence : URGENT (stock 0 ou ≤ 3 j) · MODÉRÉ (≤ 15 j) · À ANTICIPER (> 15 j).
"""

from __future__ import annotations

import io
import math
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import pandas as pd

try:  # rapidfuzz est optionnel : sans lui, seul le matching exact/CIP marche.
    from rapidfuzz import fuzz

    _RAPIDFUZZ = True
except ImportError:  # pragma: no cover - environnement sans rapidfuzz
    _RAPIDFUZZ = False

# ---------------------------------------------------------------------------
# Constantes métier
# ---------------------------------------------------------------------------

COUVERTURE_SANS_DATE_JOURS = 30   # objectif de couverture quand pas de réappro
JOURS_PAR_MOIS = 30               # convention rotation mensuelle → journalière
SEUIL_ALERTE_PEREMPTION_JOURS = 90  # DLUO à moins de ~3 mois → alerte
SEUIL_VIGILANCE_JOURS = 7         # couverture < 7 j hors rupture → vigilance
ROTATION_MIN_VIGILANCE = 5        # < 5 ventes/mois → pas de vigilance (bruit)
SEUIL_MARGE_JUSTESSE_JOURS = 3    # écarté avec < 3 j de marge → à surveiller
SEUIL_TENDANCE = 0.20             # ±20 % entre 3 mois et annuelle → ↗ / ↘

URGENT = "🔴 URGENT"
MODERE = "🟡 MODÉRÉ"
ANTICIPER = "🟢 À ANTICIPER"
_ORDRE_URGENCE = {URGENT: 0, MODERE: 1, ANTICIPER: 2}

SEUIL_MATCH = 80      # score fuzzy minimal pour accepter une correspondance
SEUIL_CERTAIN = 92    # en dessous → correspondance « incertaine », à vérifier


# ---------------------------------------------------------------------------
# Normalisation / parsing
# ---------------------------------------------------------------------------

def normaliser_libelle(libelle) -> str:
    """Majuscules, sans accents, ponctuation → espace, espaces réduits."""
    if libelle is None or (isinstance(libelle, float) and math.isnan(libelle)):
        return ""
    s = str(libelle)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normaliser_cip(cip) -> str:
    """Ne garde que les chiffres (gère les CIP lus en float : '3400930.0').

    Un CIP « 0 » (placeholder fréquent dans les exports) est traité comme
    absent — sinon deux produits distincts à CIP 0 se rapprocheraient à tort.
    """
    if cip is None or (isinstance(cip, float) and math.isnan(cip)):
        return ""
    s = str(cip).strip()
    if re.fullmatch(r"\d+\.0+", s):  # float Excel → entier
        s = s.split(".")[0]
    s = re.sub(r"\D", "", s)
    return "" if s.strip("0") == "" else s


def variantes_cip(cip: str) -> list:
    """Formes équivalentes d'un CIP pour le matching inter-fichiers.

    Les exports mélangent CIP13 et CIP7 : le CIP13 médicament français
    (13 chiffres, préfixe 3400) contient le CIP7 en positions 6-12
    (ex. 3400932300778 → 3230077, Titanoréine). On rapproche donc les deux
    formes. Les autres EAN13 (parapharmacie…) restent tels quels.
    """
    if not cip:
        return []
    formes = [cip]
    if len(cip) == 13 and cip.startswith("3400"):
        formes.append(cip[5:12])  # CIP7 embarqué dans le CIP13
    return formes


def parser_nombre(val) -> float:
    """Nombre robuste : virgule décimale française, espaces, vide → 0."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if (isinstance(val, float) and math.isnan(val)) else float(val)
    s = str(val).strip().replace(" ", "").replace(" ", "")
    if not s:
        return 0.0
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parser_date(val) -> Optional[date]:
    """Date robuste (formats français en priorité). None si illisible/vide."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in {"nan", "nat", "-", "?"}:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y",
                "%d/%m", "%d-%m"):
        try:
            d = datetime.strptime(s, fmt).date()
            if fmt in ("%d/%m", "%d-%m"):  # jour/mois sans année → année en cours
                d = d.replace(year=date.today().year)
            return d
        except ValueError:
            continue
    try:  # dernier recours : pandas, convention jour d'abord (français)
        d = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return None if pd.isna(d) else d.date()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Chargement des fichiers (.xlsx / .xls / .csv)
# ---------------------------------------------------------------------------

def charger_fichier(contenu, nom_fichier: str) -> pd.DataFrame:
    """Charge un fichier Excel/CSV en DataFrame (colonnes en str).

    ``contenu`` : bytes, chemin, ou objet fichier (upload Streamlit).
    Lève ValueError avec un message clair si le format n'est pas géré.
    """
    nom = (nom_fichier or "").lower()
    if isinstance(contenu, (str,)):  # chemin sur disque
        with open(contenu, "rb") as f:
            data = f.read()
    elif isinstance(contenu, bytes):
        data = contenu
    else:  # objet fichier (BytesIO, UploadedFile…)
        data = contenu.read()

    if nom.endswith(".csv") or nom.endswith(".txt"):
        for encodage in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                texte = data.decode(encodage)
                sep = ";" if texte.count(";") >= texte.count(",") else ","
                df = pd.read_csv(io.StringIO(texte), sep=sep)
                break
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue
        else:
            raise ValueError(f"CSV illisible : {nom_fichier}")
    elif nom.endswith(".xlsx") or nom.endswith(".xlsm"):
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
    elif nom.endswith(".xls"):
        df = pd.read_excel(io.BytesIO(data))  # xlrd requis pour les vieux .xls
    elif nom.endswith(".pdf"):
        df = _charger_pdf(data, nom_fichier)
    else:
        raise ValueError(
            f"Format non géré : {nom_fichier} (attendu .xlsx, .xls, .csv ou .pdf)")

    df.columns = [str(c).strip() for c in df.columns]
    return df


def _charger_pdf(data: bytes, nom_fichier: str) -> pd.DataFrame:
    """Extrait le tableau d'un PDF multi-pages (export type cadencier).

    Le cadencier WinPharma (en-tête « Codes produit ») est reconnu et passe
    par un parseur dédié. Sinon : la première ligne non vide sert d'en-tête,
    l'en-tête répété en haut de chaque page est éliminé. PDF sans traits de
    tableau : repli sur l'alignement du texte. PDF scanné (image) : message
    clair, préférer l'export Excel/CSV.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ValueError("Lecture PDF indisponible : lancez "
                         "« pip install pdfplumber » puis réessayez.")

    brutes: list = []
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:  # pas de traits → colonnes par alignement texte
                    table = page.extract_table({"vertical_strategy": "text",
                                                "horizontal_strategy": "text"})
                    tables = [table] if table else []
                for table in tables:
                    for brute in table:
                        valeurs = ["" if v is None else str(v).strip()
                                   for v in (brute or [])]
                        if any(valeurs):
                            brutes.append(valeurs)
    except Exception:
        raise ValueError(f"PDF illisible : {nom_fichier}")

    if not brutes:
        raise ValueError(
            f"Aucun tableau lisible dans {nom_fichier} — s'il s'agit d'un "
            "PDF scanné (image), préférez un export Excel ou CSV.")

    if any("Codes produit" in " ".join(r) for r in brutes[:40]):
        return _parser_cadencier_winpharma(brutes)

    en_tete, lignes = None, []
    for valeurs in brutes:
        if en_tete is None:
            en_tete = valeurs
        elif valeurs != en_tete:  # ignore l'en-tête répété à chaque page
            lignes.append(valeurs)
    if en_tete is None or not lignes:
        raise ValueError(
            f"Aucun tableau lisible dans {nom_fichier} — s'il s'agit d'un "
            "PDF scanné (image), préférez un export Excel ou CSV.")
    largeur = len(en_tete)  # aligne les lignes incomplètes sur l'en-tête
    lignes = [l[:largeur] + [""] * (largeur - len(l)) for l in lignes]
    return pd.DataFrame(lignes, columns=en_tete)


def _parser_cadencier_winpharma(brutes: list) -> pd.DataFrame:
    """Cadencier de stock WinPharma (PDF) → DataFrame prêt pour l'analyse.

    Format observé (4 cellules par ligne) :
      - « Codes produit » : CIP7 et CIP13 empilés (ou EAN seul) ;
      - « Nom / Formes & presentations » : libellé, parfois sur 2 lignes ;
      - « Stock » : quantité en rayon ;
      - achats/ventes : « A … » puis « V … » sur la même cellule, 12 valeurs
        mensuelles en ordre ANTI-chronologique (mois récent en premier) +
        total. On ne garde que la ligne V, remise en ordre chronologique.
    Le bandeau de la pharmacie, les en-têtes répétés par page et la ligne de
    totaux finale sont éliminés.
    """
    mois = None
    for r in brutes:
        joint = " ".join(r)
        if "Codes produit" in joint:
            trouve = re.search(r"((?:[A-Za-zéû]{3}\s+){11}[A-Za-zéû]{3})\s+Total",
                               joint)
            if trouve:
                mois = trouve.group(1).split()
            break
    if not mois:
        raise ValueError("Cadencier WinPharma : en-tête des mois introuvable.")
    colonnes_ventes = [f"Ventes {m}" for m in reversed(mois)]  # récent en DERNIER

    produits = []
    for r in brutes:
        if len(r) < 4:
            continue
        codes_cell, nom, stock_cell, achats_ventes = r[0], r[1], r[2], r[3]
        if "Codes produit" in codes_cell or "CADENCIER" in " ".join(r).upper():
            continue
        codes = re.findall(r"\d{6,}", codes_cell)
        if not codes:  # bandeau de page, ligne de totaux (« Manque: … »)…
            continue
        cip = next((c for c in codes if len(c) == 13), codes[0])
        ventes_v = re.search(r"(?:^|\n)V\s+([\d\s,.]+)", achats_ventes or "")
        ventes: list = []
        if ventes_v:
            nombres = ventes_v.group(1).split()
            if len(nombres) > 12:
                nombres = nombres[:12]  # sans la colonne Total
            ventes = list(reversed(nombres))  # → ordre chronologique
        ligne = {"Produit": " ".join((nom or "").split()), "CIP": cip,
                 "Stock": stock_cell}
        manquants = len(colonnes_ventes) - len(ventes)
        ventes = ["0"] * max(0, manquants) + ventes  # mois absents = 0 vente
        for colonne, valeur in zip(colonnes_ventes, ventes):
            ligne[colonne] = valeur
        produits.append(ligne)
    if not produits:
        raise ValueError("Cadencier WinPharma : aucune ligne produit lisible.")
    return pd.DataFrame(produits,
                        columns=["Produit", "CIP", "Stock"] + colonnes_ventes)


# ---------------------------------------------------------------------------
# Détection automatique des colonnes (proposition, à confirmer dans l'UI)
# ---------------------------------------------------------------------------

_MOTS_CLES = {
    "libelle": ["libell", "produit", "design", "article", "nom", "denomination"],
    "cip": ["cip", "code produit", "code article", "ean", "acl"],
    "stock": ["stock", "qte dispo", "quantite dispo", "disponible"],
    "date_reappro": ["reappro", "réappro", "reapprovisionnement", "retour",
                     "dispo le", "date"],
    "conditionnement": ["conditionnement", "colisage", "pcb", "unite de vente"],
    "commande_en_cours": ["commande en cours", "qte commandee", "cde en cours",
                          "en commande"],
    "peremption": ["peremption", "dluo", "date de peremption", "date limite"],
}


def _sans_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c)).lower()


def detecter_colonne(colonnes, role: str) -> Optional[str]:
    """Propose la colonne la plus probable pour un rôle donné (ou None)."""
    for mot in _MOTS_CLES.get(role, []):
        for col in colonnes:
            if _sans_accents(mot) in _sans_accents(str(col)):
                return col
    return None


def detecter_colonnes_ventes(colonnes) -> list:
    """Propose les colonnes de ventes mensuelles (mois ou mot-clé « vente »)."""
    mois = ["janv", "fevr", "mars", "avr", "mai", "juin", "juil", "aout",
            "sept", "oct", "nov", "dec"]
    trouvees = []
    for col in colonnes:
        c = _sans_accents(str(col))
        if "vente" in c or "sortie" in c or any(m in c for m in mois):
            trouvees.append(col)
    return trouvees


# ---------------------------------------------------------------------------
# Matching produit (CIP prioritaire, sinon libellé normalisé + fuzzy)
# ---------------------------------------------------------------------------

@dataclass
class Correspondance:
    """Résultat du rapprochement d'un produit avec une autre liste."""
    index: Optional[int]          # index de la ligne appariée (None = aucun)
    methode: str                  # "cip" | "exact" | "fuzzy" | "aucune"
    score: float = 100.0          # score fuzzy (100 pour cip/exact)
    incertain: bool = False       # True si score < SEUIL_CERTAIN → à vérifier


def apparier(libelle: str, cip: str, index_cip: dict, index_libelle: dict,
             libelles_norm: list) -> Correspondance:
    """Rapproche un produit d'une liste indexée (CIP > exact > fuzzy)."""
    for forme in variantes_cip(cip):  # CIP13 et CIP7 équivalents
        if forme in index_cip:
            return Correspondance(index_cip[forme], "cip")
    norme = normaliser_libelle(libelle)
    if not norme:
        return Correspondance(None, "aucune", 0.0)
    if norme in index_libelle:
        return Correspondance(index_libelle[norme], "exact")
    if _RAPIDFUZZ and libelles_norm:
        meilleur, meilleur_score = None, 0.0
        for autre_norme, idx in libelles_norm:
            s = fuzz.token_sort_ratio(norme, autre_norme)
            if s > meilleur_score:
                meilleur, meilleur_score = idx, s
        if meilleur is not None and meilleur_score >= SEUIL_MATCH:
            return Correspondance(meilleur, "fuzzy", meilleur_score,
                                  incertain=meilleur_score < SEUIL_CERTAIN)
    return Correspondance(None, "aucune", 0.0)


def _indexer(df: pd.DataFrame, col_libelle: str, col_cip: Optional[str]):
    """Construit les index (CIP → ligne, libellé normalisé → ligne)."""
    index_cip, index_libelle, libelles_norm = {}, {}, []
    for idx, row in df.iterrows():
        if col_cip and col_cip in df.columns:
            for forme in variantes_cip(normaliser_cip(row[col_cip])):
                if forme not in index_cip:  # CIP13 ET CIP7 indexés
                    index_cip[forme] = idx
        norme = normaliser_libelle(row[col_libelle])
        if norme:
            if norme not in index_libelle:
                index_libelle[norme] = idx
            libelles_norm.append((norme, idx))
    return index_cip, index_libelle, libelles_norm


# ---------------------------------------------------------------------------
# Calculs élémentaires (étapes 2, 3, 5, 6) — unitairement testables
# ---------------------------------------------------------------------------

def calculer_rotation_mensuelle(ventes: list, periode: str = "annuelle") -> float:
    """Moyenne mensuelle des ventes.

    ``ventes`` : valeurs mensuelles en ordre CHRONOLOGIQUE (la plus récente en
    dernier). ``periode`` : "annuelle" (moyenne de tout) ou "3mois" (moyenne
    des 3 dernières valeurs).
    """
    valeurs = [parser_nombre(v) for v in ventes]
    if not valeurs:
        return 0.0
    if periode == "3mois":
        valeurs = valeurs[-3:]
    return sum(valeurs) / len(valeurs)


def calculer_stock_jours(stock_actuel: float, rotation_mensuelle: float) -> float:
    """Couverture actuelle en jours. Stock 0 → 0 ; rotation nulle → +inf."""
    if stock_actuel <= 0:
        return 0.0
    rotation_journaliere = rotation_mensuelle / JOURS_PAR_MOIS
    if rotation_journaliere <= 0:
        return math.inf
    return stock_actuel / rotation_journaliere


def doit_apparaitre(stock_jours: float,
                    jours_avant_reappro: Optional[float]) -> bool:
    """Étape 3 — règle d'apparition STRICTE, sans buffer.

    - Avec date de réappro : apparaît ssi stock_jours < jours_avant_reappro.
    - Sans date : apparaît ssi stock_jours < 30.
    """
    if jours_avant_reappro is not None:
        return stock_jours < jours_avant_reappro
    return stock_jours < COUVERTURE_SANS_DATE_JOURS


def classer_urgence(stock_actuel: float, stock_jours: float) -> str:
    """Étape 6 — URGENT (stock 0 ou ≤ 3 j) · MODÉRÉ (≤ 15 j) · À ANTICIPER."""
    if stock_actuel <= 0 or stock_jours <= 3:
        return URGENT
    if stock_jours <= 15:
        return MODERE
    return ANTICIPER


def quantite_a_commander(rotation_mensuelle: float,
                         couverture_cible_jours: float,
                         stock_actuel: float,
                         conditionnement: Optional[float] = None) -> int:
    """Étape 5 — Cmd = arrondi_sup(rot_j × couverture − stock), minimum 1.

    Arrondi au multiple du conditionnement si fourni (> 1).
    """
    rotation_journaliere = rotation_mensuelle / JOURS_PAR_MOIS
    qte_cible = rotation_journaliere * couverture_cible_jours
    cmd = max(1, math.ceil(qte_cible - stock_actuel))
    if conditionnement and conditionnement > 1:
        cmd = int(math.ceil(cmd / conditionnement) * conditionnement)
    return cmd


def calculer_tendance(ventes: list, seuil: float = SEUIL_TENDANCE) -> str:
    """Tendance de la demande, en ordre chronologique (récent en dernier).

    - ≥ 4 mois de recul : moyenne des 3 derniers mois vs moyenne globale ;
    - 2-3 mois (cadenciers courts, très fréquents) : dernier mois vs moyenne
      des mois précédents — moins robuste mais la colonne reste vivante ;
    - < 2 mois ou demande nulle : « → stable » (rien à comparer).
    Renvoie « ↗ hausse » / « ↘ baisse » / « → stable » (seuil ±20 %).
    """
    if len(ventes) < 2:
        return "→ stable"
    if len(ventes) >= 4:
        reference = calculer_rotation_mensuelle(ventes, "annuelle")
        recente = calculer_rotation_mensuelle(ventes, "3mois")
    else:
        reference = calculer_rotation_mensuelle(ventes[:-1], "annuelle")
        recente = parser_nombre(ventes[-1])
    if reference <= 0:
        return "→ stable"
    ecart = (recente - reference) / reference
    if ecart >= seuil:
        return "↗ hausse"
    if ecart <= -seuil:
        return "↘ baisse"
    return "→ stable"


def compter_reports_reappro(produit: str, historique,
                            date_reappro_jour: Optional[date] = None) -> int:
    """Nombre de fois où la date de réappro annoncée pour ``produit`` a été
    REPOUSSÉE d'une analyse à l'autre (fournisseur peu fiable sur ce produit).

    ``date_reappro_jour`` : date annoncée AUJOURD'HUI, comparée en dernier —
    un glissement entre hier et aujourd'hui compte donc dès aujourd'hui.
    L'historique doit contenir les colonnes 'Date analyse', 'Produit' et
    'Date réappro' (les anciens historiques sans cette colonne renvoient 0).
    """
    annonces: list = []
    if (historique is not None and not historique.empty
            and "Date réappro" in historique.columns):
        sous = historique[historique["Produit"] == produit].copy()
        if not sous.empty:
            sous["_date"] = pd.to_datetime(sous["Date analyse"], errors="coerce")
            sous = sous.sort_values("_date")
            annonces = [parser_date(v) for v in sous["Date réappro"]]
    return _compter_reports(annonces, date_reappro_jour)


def _compter_reports(annonces: list, date_reappro_jour: Optional[date]) -> int:
    """Compte les glissements dans une suite chronologique de dates annoncées
    (None = pas de date ce jour-là, ignoré), la date du jour en dernier."""
    reports, derniere = 0, None
    for d in list(annonces) + [date_reappro_jour]:
        if d is None:
            continue
        if derniere is not None and d > derniere:
            reports += 1
        derniere = d
    return reports


def rotation_possiblement_sous_estimee(ventes: list) -> bool:
    """Indice de rupture passée : au moins un mois à 0 vente alors que
    d'autres mois montrent des ventes — la rotation calculée est alors
    probablement sous-estimée (le produit était en rupture, pas sans
    demande). Toutes les valeurs à 0 = pas de demande réelle → non signalé.
    """
    valeurs = [parser_nombre(v) for v in ventes]
    if not valeurs or all(v <= 0 for v in valeurs):
        return False
    return any(v <= 0 for v in valeurs)


def compter_occurrences_historique(produit: str, historique: pd.DataFrame,
                                   avant_date: date) -> int:
    """Nombre d'analyses antérieures à ``avant_date`` où ``produit`` était
    déjà signalé dans l'historique (colonnes 'Date analyse' / 'Produit',
    persisté par l'interface — voir app.py). Module pur : aucune I/O ici.
    """
    if historique is None or historique.empty:
        return 0
    sous = historique[historique["Produit"] == produit]
    dates = pd.to_datetime(sous["Date analyse"], errors="coerce").dt.date
    return int((dates < avant_date).sum())


def comparer_a_analyse_precedente(produits_jour, historique: pd.DataFrame,
                                  date_analyse: date):
    """Suivi quotidien : compare les produits signalés aujourd'hui à la
    DERNIÈRE analyse antérieure à ``date_analyse``.

    Renvoie ``(date_precedente, nouveaux, resolus)`` :
      - date_precedente : date de l'analyse de référence (None si aucune —
        première analyse, tout est « nouveau ») ;
      - nouveaux : produits du jour absents de l'analyse précédente ;
      - resolus  : produits de l'analyse précédente absents aujourd'hui.
    """
    produits_jour = list(produits_jour)
    if historique is None or historique.empty:
        return None, produits_jour, []
    dates = pd.to_datetime(historique["Date analyse"], errors="coerce").dt.date
    anterieures = {d for d in dates if pd.notna(d) and d < date_analyse}
    if not anterieures:
        return None, produits_jour, []
    precedente = max(anterieures)
    produits_precedents = set(historique.loc[dates == precedente, "Produit"])
    nouveaux = [p for p in produits_jour if p not in produits_precedents]
    resolus = sorted(produits_precedents - set(produits_jour))
    return precedente, nouveaux, resolus


# ---------------------------------------------------------------------------
# Analyse complète
# ---------------------------------------------------------------------------

@dataclass
class ResultatAnalyse:
    """Sortie de l'analyse : les onglets de décision + résumé + alertes."""
    onglet1: pd.DataFrame           # à commander chez UNIPHARMA
    onglet2: pd.DataFrame           # rupture chez les deux → pas de solution
    onglet3: pd.DataFrame           # traçabilité complète
    resume: dict = field(default_factory=dict)
    alertes: list = field(default_factory=list)          # messages pour l'UI
    matchs_incertains: list = field(default_factory=list)  # à vérifier à la main
    vigilance: pd.DataFrame = field(default_factory=pd.DataFrame)
    # ^ anticipation : produits hors rupture GPNC dont le stock s'épuise
    ecartes_justesse: pd.DataFrame = field(default_factory=pd.DataFrame)
    # ^ écartés par la règle stricte mais avec très peu de marge


COLONNES_ONGLET1 = ["Urgence", "Produit", "Stock actuel", "Commande en cours",
                    "Rotation/mois", "Tendance", "Fiabilité rotation",
                    "Stock (jours)", "Date réappro GPNC", "Jours avant réappro",
                    "Péremption", "Qté à commander (Cmd)", "Commentaire"]
COLONNES_VIGILANCE = ["Produit", "Stock actuel", "Commande en cours",
                      "Rotation/mois", "Tendance", "Stock (jours)", "Conseil"]
COLONNES_JUSTESSE = ["Produit", "Stock actuel", "Rotation/mois",
                     "Stock (jours)", "Date réappro GPNC",
                     "Jours avant réappro", "Marge (jours)", "Commentaire"]
COLONNES_ONGLET2 = ["Produit", "Stock actuel", "Rotation/mois", "Stock (jours)",
                    "Date réappro GPNC", "Péremption", "Commentaire"]
COLONNES_ONGLET3 = ["Produit", "Vendu (O/N)", "Stock actuel", "Commande en cours",
                    "Rotation/mois", "Fiabilité rotation", "Stock (jours)",
                    "Date réappro", "Jours avant réappro", "Péremption",
                    "Dispo UNIPHARMA (O/N)", "Décision", "Onglet", "Motif"]


def analyser(cadencier: pd.DataFrame,
             ruptures_gpnc: pd.DataFrame,
             ruptures_unipharma: pd.DataFrame,
             mapping: dict,
             date_analyse: date,
             periode: str = "annuelle",
             historique: Optional[pd.DataFrame] = None,
             seuil_vigilance_jours: float = SEUIL_VIGILANCE_JOURS,
             rotation_min_vigilance: float = ROTATION_MIN_VIGILANCE,
             seuil_marge_jours: float = SEUIL_MARGE_JUSTESSE_JOURS,
             delai_livraison_jours: float = 0,
             rotation_prudente: bool = False) -> ResultatAnalyse:
    """Croise les 3 fichiers et produit les onglets de décision.

    Paramètres d'anticipation (tous facultatifs, comportement historique
    par défaut) :
      - historique : analyses passées (suivi des réappros repoussées et des
        ruptures longues aux ventes écrasées à 0) ;
      - seuil_vigilance_jours : produits du cadencier HORS rupture GPNC dont
        la couverture passe sous ce seuil → onglet Vigilance ;
      - rotation_min_vigilance : plancher de ventes/mois pour la vigilance
        (écarte le bruit des produits à très faible rotation) ;
      - seuil_marge_jours : produits écartés par la règle stricte avec moins
        de cette marge → onglet Écartés de justesse ;
      - delai_livraison_jours : délai de livraison UNIPHARMA ajouté à la
        couverture cible du calcul de Cmd (pas à la règle d'apparition) ;
      - rotation_prudente : retient max(annuelle, 3 mois) par produit.

    ``mapping`` décrit les colonnes de chaque fichier :
      {
        "cadencier":  {"libelle": str, "cip": str|None, "stock": str,
                        "ventes": [str, ...],  # ordre chrono, récent en dernier
                        "conditionnement": str|None,
                        "commande_en_cours": str|None,  # qté déjà commandée
                        "peremption": str|None},        # DLUO la plus proche
        "gpnc":       {"libelle": str, "cip": str|None, "date_reappro": str|None},
        "unipharma":  {"libelle": str, "cip": str|None},
      }
    """
    m_cad, m_gpnc, m_uni = mapping["cadencier"], mapping["gpnc"], mapping["unipharma"]
    alertes: list = []
    matchs_incertains: list = []

    # Index du cadencier et des ruptures UNIPHARMA pour le rapprochement.
    idx_cip_cad, idx_lib_cad, libs_cad = _indexer(
        cadencier, m_cad["libelle"], m_cad.get("cip"))
    idx_cip_uni, idx_lib_uni, libs_uni = _indexer(
        ruptures_unipharma, m_uni["libelle"], m_uni.get("cip"))

    if not _RAPIDFUZZ:
        alertes.append("rapidfuzz non installé : matching exact/CIP uniquement "
                       "(pip install rapidfuzz recommandé).")

    def _rotation(ventes: list) -> float:
        """Rotation retenue : période choisie, ou la plus élevée des deux
        moyennes (annuelle / 3 mois) en mode prudent — un produit en
        croissance n'est alors jamais sous-couvert."""
        if rotation_prudente:
            return max(calculer_rotation_mensuelle(ventes, "annuelle"),
                       calculer_rotation_mensuelle(ventes, "3mois"))
        return calculer_rotation_mensuelle(ventes, periode)

    def _extraire_cadencier(ligne_cad):
        """Valeurs numériques d'une ligne du cadencier — partagé entre la
        boucle des ruptures GPNC et le balayage Vigilance, pour que les deux
        calculent la couverture avec exactement les mêmes règles."""
        stock = parser_nombre(ligne_cad[m_cad["stock"]])
        en_cours = (parser_nombre(ligne_cad[m_cad["commande_en_cours"]])
                    if m_cad.get("commande_en_cours") else 0.0)
        ventes = [ligne_cad[c] for c in m_cad["ventes"]
                  if c in cadencier.columns]
        return stock, en_cours, ventes, _rotation(ventes)

    # Historique pré-groupé par produit (une seule passe de parsing/tri au
    # lieu d'un filtre complet du DataFrame par produit signalé).
    annonces_reappro: dict = {}
    if (historique is not None and not historique.empty
            and "Date réappro" in historique.columns):
        h = historique.copy()
        h["_date"] = pd.to_datetime(h["Date analyse"], errors="coerce")
        for produit, groupe in h.sort_values("_date").groupby("Produit",
                                                              sort=False):
            annonces_reappro[produit] = [parser_date(v)
                                         for v in groupe["Date réappro"]]

    lignes1, lignes2, lignes3 = [], [], []
    lignes_justesse: list = []
    indices_cadencier_traites: set = set()  # pour l'onglet Vigilance

    for _, ligne_gpnc in ruptures_gpnc.iterrows():
        produit_gpnc = str(ligne_gpnc[m_gpnc["libelle"]]).strip()
        if not produit_gpnc or produit_gpnc.lower() == "nan":
            continue
        cip_gpnc = normaliser_cip(ligne_gpnc[m_gpnc["cip"]]) if m_gpnc.get("cip") else ""

        # --- Rapprochement avec le cadencier -------------------------------
        corr = apparier(produit_gpnc, cip_gpnc, idx_cip_cad, idx_lib_cad, libs_cad)
        if corr.incertain:
            matchs_incertains.append({
                "Produit (ruptures GPNC)": produit_gpnc,
                "Rapproché de (cadencier)":
                    str(cadencier.loc[corr.index, m_cad["libelle"]]),
                "Score": round(corr.score, 1), "Fichier": "cadencier",
            })

        # --- Date de réappro ------------------------------------------------
        date_reappro = (parser_date(ligne_gpnc[m_gpnc["date_reappro"]])
                        if m_gpnc.get("date_reappro") else None)
        jours_avant = None
        if date_reappro is not None:
            jours_avant = (date_reappro - date_analyse).days
            if jours_avant < 0:  # réappro passée → traité comme sans date
                alertes.append(f"{produit_gpnc} : date de réappro dépassée "
                               f"({date_reappro:%d/%m/%Y}) — traité comme sans date.")
                date_reappro, jours_avant = None, None

        base3 = {
            "Produit": produit_gpnc,
            "Date réappro": f"{date_reappro:%d/%m/%Y}" if date_reappro else "",
            "Jours avant réappro": jours_avant if jours_avant is not None else "",
        }

        # --- Étape 1 : périmètre (vendu, rotation > 0) ----------------------
        if corr.index is None:
            lignes3.append({**base3, "Vendu (O/N)": "N", "Stock actuel": "",
                            "Commande en cours": "", "Rotation/mois": "",
                            "Fiabilité rotation": "", "Stock (jours)": "",
                            "Péremption": "", "Dispo UNIPHARMA (O/N)": "",
                            "Décision": "Écarté", "Onglet": "—",
                            "Motif": "Absent du cadencier (non vendu)"})
            continue

        ligne_cad = cadencier.loc[corr.index]
        indices_cadencier_traites.add(corr.index)
        stock, en_cours, ventes, rotation = _extraire_cadencier(ligne_cad)

        # Commande en cours : évite de recommander ce qui arrive déjà.
        stock_effectif = stock + en_cours
        affiche_en_cours = en_cours if m_cad.get("commande_en_cours") else ""

        # --- Péremption (DLUO) : alerte informative, n'écarte pas le produit
        date_peremption = (parser_date(ligne_cad[m_cad["peremption"]])
                           if m_cad.get("peremption") else None)
        affiche_peremption = ""
        if date_peremption is not None:
            affiche_peremption = f"{date_peremption:%d/%m/%Y}"
            jours_avant_peremption = (date_peremption - date_analyse).days
            if 0 <= jours_avant_peremption <= SEUIL_ALERTE_PEREMPTION_JOURS:
                alertes.append(
                    f"{produit_gpnc} : péremption proche ({affiche_peremption}, "
                    f"dans {jours_avant_peremption} j) — vérifier le stock "
                    "avant de commander davantage.")

        tendance = calculer_tendance(ventes)
        rotation_douteuse = rotation_possiblement_sous_estimee(ventes)
        affiche_fiabilite = ("⚠️ rupture passée possible" if rotation_douteuse
                             else "OK")
        if rotation <= 0:
            # Rupture LONGUE : ventes écrasées à 0 sur toute la période, mais
            # le produit était déjà signalé → ne pas l'écarter en silence.
            deja_signale = compter_occurrences_historique(
                produit_gpnc, historique, date_analyse)
            if deja_signale > 0:
                alertes.append(
                    f"{produit_gpnc} : ventes à 0 sur toute la période mais "
                    f"déjà signalé {deja_signale} fois — rupture longue "
                    "probable, rotation incalculable ; vérifier manuellement "
                    "(dépannage UNIPHARMA possible).")
                decision, motif = "À vérifier", (
                    f"Rotation nulle mais déjà signalé {deja_signale} fois "
                    "(rupture longue probable)")
            else:
                decision, motif = "Écarté", "Rotation nulle (produit non vendu)"
            lignes3.append({**base3, "Vendu (O/N)": "N", "Stock actuel": stock,
                            "Commande en cours": affiche_en_cours,
                            "Rotation/mois": 0, "Fiabilité rotation": "",
                            "Stock (jours)": "", "Péremption": affiche_peremption,
                            "Dispo UNIPHARMA (O/N)": "", "Décision": decision,
                            "Onglet": "—", "Motif": motif})
            continue

        # --- Étapes 2-3 : stock en jours + règle d'apparition ---------------
        # Le stock EFFECTIF (physique + commandes en cours) sert de base au
        # calcul : une commande déjà partie ne doit pas être recommandée.
        stock_jours = calculer_stock_jours(stock_effectif, rotation)
        base3.update({"Vendu (O/N)": "O", "Stock actuel": stock,
                      "Commande en cours": affiche_en_cours,
                      "Rotation/mois": round(rotation, 1),
                      "Fiabilité rotation": affiche_fiabilite,
                      "Stock (jours)": round(stock_jours, 1),
                      "Péremption": affiche_peremption})

        if not doit_apparaitre(stock_jours, jours_avant):
            motif = (f"Stock ({stock_jours:.0f} j) couvre jusqu'à la réappro "
                     f"({jours_avant} j)" if jours_avant is not None
                     else f"Stock ({stock_jours:.0f} j) ≥ 30 j de couverture")
            # Écarté de JUSTESSE : la règle stricte tient, mais avec si peu
            # de marge qu'un glissement de réappro suffirait → visible.
            marge = stock_jours - (jours_avant if jours_avant is not None
                                   else COUVERTURE_SANS_DATE_JOURS)
            if marge < seuil_marge_jours:
                lignes_justesse.append({
                    "Produit": produit_gpnc, "Stock actuel": stock,
                    "Rotation/mois": round(rotation, 1),
                    "Stock (jours)": round(stock_jours, 1),
                    "Date réappro GPNC": base3["Date réappro"],
                    "Jours avant réappro": (jours_avant
                                            if jours_avant is not None else ""),
                    "Marge (jours)": round(marge, 1),
                    "Commentaire": ("Écarté par la règle stricte mais marge "
                                    "faible — si la réappro glisse, rupture "
                                    "sèche. Surveiller / dépanner au besoin."),
                })
                motif += f" — de justesse ({marge:.1f} j de marge)"
            lignes3.append({**base3, "Dispo UNIPHARMA (O/N)": "",
                            "Décision": "Écarté", "Onglet": "—", "Motif": motif})
            continue

        # --- Étape 4 : dépannage UNIPHARMA ----------------------------------
        corr_uni = apparier(produit_gpnc, cip_gpnc, idx_cip_uni, idx_lib_uni, libs_uni)
        if corr_uni.incertain:
            matchs_incertains.append({
                "Produit (ruptures GPNC)": produit_gpnc,
                "Rapproché de (cadencier)":
                    str(ruptures_unipharma.loc[corr_uni.index, m_uni["libelle"]]),
                "Score": round(corr_uni.score, 1), "Fichier": "ruptures UNIPHARMA",
            })
        rupture_uni = corr_uni.index is not None
        urgence = classer_urgence(stock_effectif, stock_jours)

        if rupture_uni:  # rupture chez les DEUX → pas de solution
            lignes2.append({
                "Produit": produit_gpnc, "Stock actuel": stock,
                "Rotation/mois": round(rotation, 1),
                "Stock (jours)": round(stock_jours, 1),
                "Date réappro GPNC": base3["Date réappro"],
                "Péremption": affiche_peremption,
                "Commentaire": ("Anticiper l'information patient ; contacter "
                                "GPNC pour confirmer la date de réappro."),
                "_rotation": rotation, "_stock": stock,
            })
            lignes3.append({**base3, "Dispo UNIPHARMA (O/N)": "N",
                            "Décision": "Retenu", "Onglet": "Onglet 2",
                            "Motif": "Rupture GPNC + UNIPHARMA (pas de solution)"})
            continue

        # --- Étape 5 : quantité à commander ---------------------------------
        # Le délai de livraison UNIPHARMA s'ajoute à la couverture cible :
        # les boîtes commandées aujourd'hui n'arrivent pas aujourd'hui.
        couverture_cible = (jours_avant if jours_avant is not None
                            else COUVERTURE_SANS_DATE_JOURS) + delai_livraison_jours
        conditionnement = None
        if m_cad.get("conditionnement"):
            c = parser_nombre(ligne_cad[m_cad["conditionnement"]])
            conditionnement = c if c > 1 else None
        cmd = quantite_a_commander(rotation, couverture_cible, stock_effectif,
                                   conditionnement)

        commentaire = ("Dépannage jusqu'à la réappro GPNC" if jours_avant is not None
                       else "Pas de date de réappro → objectif 30 j de couverture")
        if en_cours:
            commentaire += f" · {en_cours:g} déjà en commande (déduit du calcul)"
        reports = _compter_reports(annonces_reappro.get(produit_gpnc, []),
                                   date_reappro)
        if reports:
            commentaire += (f" · ⚠️ réappro déjà repoussée {reports} fois "
                            "(date peu fiable)")
            alertes.append(f"{produit_gpnc} : la date de réappro GPNC a déjà "
                           f"été repoussée {reports} fois — ne pas compter "
                           "dessus, privilégier le dépannage.")
        lignes1.append({
            "Urgence": urgence, "Produit": produit_gpnc, "Stock actuel": stock,
            "Commande en cours": affiche_en_cours,
            "Rotation/mois": round(rotation, 1),
            "Tendance": tendance,
            "Fiabilité rotation": affiche_fiabilite,
            "Stock (jours)": round(stock_jours, 1),
            "Date réappro GPNC": base3["Date réappro"],
            "Jours avant réappro": jours_avant if jours_avant is not None else "",
            "Péremption": affiche_peremption,
            "Qté à commander (Cmd)": cmd, "Commentaire": commentaire,
            "_stock_jours": stock_jours,
        })
        lignes3.append({**base3, "Dispo UNIPHARMA (O/N)": "O",
                        "Décision": "Retenu", "Onglet": "Onglet 1",
                        "Motif": f"À commander chez UNIPHARMA ({urgence})"})

    # --- Vigilance : anticiper les ruptures de VOTRE stock -------------------
    # Produits du cadencier HORS liste de ruptures GPNC dont la couverture
    # passe sous le seuil : la rupture en rayon arrive, autant commander
    # chez GPNC avant qu'elle se produise.
    lignes_vigilance: list = []
    for idx, ligne_cad in cadencier.iterrows():
        if idx in indices_cadencier_traites:
            continue  # déjà couvert par l'analyse des ruptures GPNC
        stock, en_cours, ventes, rotation = _extraire_cadencier(ligne_cad)
        if rotation < rotation_min_vigilance or rotation <= 0:
            continue  # rotation trop faible : bruit, rien à anticiper
        stock_jours = calculer_stock_jours(stock + en_cours, rotation)
        if stock_jours >= seuil_vigilance_jours:
            continue
        lignes_vigilance.append({
            "Produit": str(ligne_cad[m_cad["libelle"]]).strip(),
            "Stock actuel": stock,
            "Commande en cours": (en_cours if m_cad.get("commande_en_cours")
                                  else ""),
            "Rotation/mois": round(rotation, 1),
            "Tendance": calculer_tendance(ventes),
            "Stock (jours)": round(stock_jours, 1),
            "Conseil": ("Hors rupture GPNC — commander chez GPNC avant la "
                        "rupture en rayon."),
            "_stock_jours": stock_jours,
        })

    # --- Tris et mise en forme des onglets ----------------------------------
    df1 = pd.DataFrame(lignes1)
    if not df1.empty:
        df1["_ordre"] = df1["Urgence"].map(_ORDRE_URGENCE)
        df1 = (df1.sort_values(["_ordre", "_stock_jours"])
                  .drop(columns=["_ordre", "_stock_jours"]))
    df1 = df1.reindex(columns=COLONNES_ONGLET1)

    df2 = pd.DataFrame(lignes2)
    if not df2.empty:  # criticité : stock 0 d'abord, puis fort volume
        df2["_stock0"] = (df2["_stock"] <= 0).astype(int)
        df2 = (df2.sort_values(["_stock0", "_rotation"], ascending=[False, False])
                  .drop(columns=["_stock0", "_rotation", "_stock"]))
    df2 = df2.reindex(columns=COLONNES_ONGLET2)

    df3 = pd.DataFrame(lignes3).reindex(columns=COLONNES_ONGLET3)

    df_vigilance = pd.DataFrame(lignes_vigilance)
    if not df_vigilance.empty:  # le plus critique puis le plus gros vendeur
        df_vigilance = (df_vigilance
                        .sort_values(["_stock_jours", "Rotation/mois"],
                                     ascending=[True, False])
                        .drop(columns=["_stock_jours"]))
    df_vigilance = df_vigilance.reindex(columns=COLONNES_VIGILANCE)

    df_justesse = pd.DataFrame(lignes_justesse)
    if not df_justesse.empty:  # la marge la plus faible en premier
        df_justesse = df_justesse.sort_values("Marge (jours)")
    df_justesse = df_justesse.reindex(columns=COLONNES_JUSTESSE)

    nb_urgents = int((df1["Urgence"] == URGENT).sum()) if not df1.empty else 0
    nb_moderes = int((df1["Urgence"] == MODERE).sum()) if not df1.empty else 0
    nb_anticiper = int((df1["Urgence"] == ANTICIPER).sum()) if not df1.empty else 0
    nb_rotation_douteuse = (int((df1["Fiabilité rotation"]
                                == "⚠️ rupture passée possible").sum())
                           if not df1.empty else 0)
    nb_peremption_proche = len([a for a in alertes if "péremption proche" in a])
    resume = {
        "ruptures_gpnc": len(ruptures_gpnc),
        "analyses": len(df3),
        "vendus": int((df3["Vendu (O/N)"] == "O").sum()) if not df3.empty else 0,
        "a_commander": len(df1),
        "sans_solution": len(df2),
        "urgents": nb_urgents, "moderes": nb_moderes, "anticiper": nb_anticiper,
        "rotation_douteuse": nb_rotation_douteuse,
        "peremption_proche": nb_peremption_proche,
        "vigilance": len(df_vigilance),
        "justesse": len(df_justesse),
    }
    return ResultatAnalyse(df1, df2, df3, resume, alertes, matchs_incertains,
                           vigilance=df_vigilance,
                           ecartes_justesse=df_justesse)


# ---------------------------------------------------------------------------
# Export Excel (3 onglets, mise en forme)
# ---------------------------------------------------------------------------

_COULEURS_URGENCE = {URGENT: "F8CBAD", MODERE: "FFE699", ANTICIPER: "C6EFCE"}


def exporter_excel(resultat: ResultatAnalyse) -> bytes:
    """Génère le classeur Excel (en-têtes gras figés, largeurs auto, couleurs)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()
    onglets = [("À commander UNIPHARMA", resultat.onglet1),
               ("Rupture GPNC+UNIPHARMA", resultat.onglet2),
               ("Vigilance stock", resultat.vigilance),
               ("Écartés de justesse", resultat.ecartes_justesse),
               ("Analyse complète", resultat.onglet3)]
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nom, df in onglets:
            df.to_excel(writer, sheet_name=nom, index=False)
            ws = writer.sheets[nom]
            ws.freeze_panes = "A2"
            for cellule in ws[1]:  # en-tête gras sur fond gris clair
                cellule.font = Font(bold=True)
                cellule.fill = PatternFill("solid", fgColor="D9D9D9")
                cellule.alignment = Alignment(vertical="center")
            for j, col in enumerate(df.columns, start=1):  # largeurs auto
                largeur = max([len(str(col))] +
                              [len(str(v)) for v in df[col].head(200)] or [10])
                ws.column_dimensions[get_column_letter(j)].width = min(largeur + 3, 45)
            if "Urgence" in df.columns:  # code couleur des lignes par urgence
                pos = list(df.columns).index("Urgence")
                for i, urgence in enumerate(df["Urgence"], start=2):
                    couleur = _COULEURS_URGENCE.get(urgence)
                    if couleur:
                        for cellule in ws[i]:
                            cellule.fill = PatternFill("solid", fgColor=couleur)
                        _ = pos  # la ligne entière est teintée, pas juste la cellule
    return buffer.getvalue()


def nom_fichier_sortie(date_analyse: date) -> str:
    """Nom conventionnel du fichier généré."""
    return f"commande_ruptures_{date_analyse:%Y-%m-%d}.xlsx"
